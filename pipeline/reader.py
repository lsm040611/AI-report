"""엑셀 원본 읽기. 서식을 잃지 않는 것이 이 모듈의 존재 이유.

pandas.read_excel 은 셀 서식을 전부 버리므로 R-05(강조 서식 -> 의미)를
구현할 수 없다. openpyxl 의 rich_text 경로로 셀 안의 조각(run)별
색·굵기·밑줄을 읽어 계약의 emphasis 어휘로 변환한다.

R-05 매핑
    빨강            -> issue_expression      (문제 표현)
    굵게 + 밑줄     -> corrected_expression  (교정 표현)
    굵게            -> key_concept           (핵심 개념)
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field
from typing import Any, List, Optional

from openpyxl import load_workbook

try:                                    # openpyxl 3.1+
    from openpyxl.cell.rich_text import CellRichText
except ImportError:                     # pragma: no cover
    CellRichText = ()                   # type: ignore[assignment]


@dataclass
class Cell:
    coord: str
    value: Any
    runs: List[dict] = field(default_factory=list)   # [{"text","emphasis"}]

    @property
    def text(self) -> str:
        if self.runs:
            return "".join(r["text"] for r in self.runs)
        return "" if self.value is None else str(self.value)


@dataclass
class Sheet:
    name: str
    grid: List[List[Cell]]

    def row(self, i: int) -> List[Cell]:
        return self.grid[i] if 0 <= i < len(self.grid) else []

    @property
    def height(self) -> int:
        return len(self.grid)

    @property
    def width(self) -> int:
        return max((len(r) for r in self.grid), default=0)


def read_workbook(path: str) -> List[Sheet]:
    """원본은 절대 수정하지 않는다. 읽기 전용으로만 연다."""
    try:
        wb = load_workbook(path, data_only=True, rich_text=True)
    except TypeError:
        # openpyxl 3.0 이하 — 부분 서식을 읽을 수 없다. 셀 전체 서식만 살린다.
        wb = load_workbook(path, data_only=True)

    sheets = []
    for ws in wb.worksheets:
        grid = []
        for row in ws.iter_rows():
            grid.append([Cell(c.coordinate, _plain(c.value), _to_runs(c)) for c in row])
        sheets.append(Sheet(ws.title, grid))
    wb.close()
    return sheets


def _plain(value):
    """`.value` 에는 저장 가능한 형태만 남긴다.

    서식이 섞인 셀의 원래 값은 openpyxl 의 CellRichText — 파이썬 객체다.
    이게 카드에 그대로 실리면 DB(JSON 컬럼)에 저장할 때 터진다.
    서식 정보는 `.runs` 가 이미 갖고 있으므로, `.value` 는 평문이면 충분하다.
    """
    if CellRichText and isinstance(value, CellRichText):
        return "".join(b if isinstance(b, str) else (getattr(b, "text", "") or "")
                       for b in value)
    return value


def _to_runs(cell) -> List[dict]:
    """셀 -> runs. 서식이 전혀 없으면 빈 리스트(= 단일 평문)를 준다."""
    value = cell.value
    if value is None:
        return []

    if CellRichText and isinstance(value, CellRichText):
        runs = []
        for block in value:
            if isinstance(block, str):
                if block:
                    runs.append({"text": block, "emphasis": None})
                continue
            text = getattr(block, "text", "") or ""
            if not text:
                continue
            font = getattr(block, "font", None)
            runs.append({"text": text, "emphasis": _emphasis(font),
                         "format": _format_desc(font)})
        return _merge(runs)

    if isinstance(value, str):
        emph = _emphasis(cell.font)          # 셀 전체에 걸린 서식
        return ([{"text": value, "emphasis": emph,
                  "format": _format_desc(cell.font)}] if emph else [])

    return []


def _emphasis(font) -> Optional[str]:
    """서식 → 후보 구간. **뜻이 아니라 '여기를 보라'는 표시만 읽는다.**

    색 배합은 파일마다 다르다. 배포 샘플은 빨강을 고칠 표현에 썼지만, 다른
    강사는 보라를 쓰고 밑줄만으로 핵심을 짚기도 한다. 그래서 특정 색을
    찾는 대신 **기본 서식에서 벗어난 곳이면 모두 후보로 잡는다.**
    실제로 색만 바꾼 평가지에서 표시 세 개 중 두 개가 통째로 사라졌다.

    뜻은 R-05 가 문장을 읽어 다시 판정한다. 여기서 굳이 나누는 것은
    R-05 를 돌리지 못하는 목 모드에서도 최소한의 구분을 남기기 위해서다.
    확신할 수 없는 것은 전부 중립(key_concept)으로 둔다 — 고칠 표현을
    권장 표현으로 뒤집어 보여 주는 것이 아무 표시도 없는 것보다 나쁘다.
    """
    if font is None:
        return None
    bold = bool(getattr(font, "b", False) or getattr(font, "bold", False))
    under = bool(getattr(font, "u", None) or getattr(font, "underline", None))
    italic = bool(getattr(font, "i", False) or getattr(font, "italic", False))
    strike = bool(getattr(font, "strike", False))
    colored = _has_color(getattr(font, "color", None))

    if strike or _is_red(getattr(font, "color", None)):
        return "issue_expression"          # 붉은색·취소선은 관례가 뚜렷하다
    if bold and under:
        return "corrected_expression"      # 계약이 정한 교정 표현 표기
    if bold or under or italic or colored:
        return "key_concept"               # 표시는 있으나 뜻은 R-05 가 정한다
    return None


def _format_desc(font) -> Optional[str]:
    """강사가 실제로 무엇을 칠했는지 사람 말로 적는다.

    R-05 가 뜻을 판정할 때 이 설명을 읽는다. '굵게' 로 뭉뚱그리면
    보라색으로 칠한 것을 굵게 칠했다고 잘못 알려 주게 된다.
    """
    if font is None:
        return None
    parts = []
    if getattr(font, "b", False) or getattr(font, "bold", False):
        parts.append("굵게")
    if getattr(font, "u", None) or getattr(font, "underline", None):
        parts.append("밑줄")
    if getattr(font, "i", False) or getattr(font, "italic", False):
        parts.append("기울임")
    if getattr(font, "strike", False):
        parts.append("취소선")
    rgb = _rgb_of(getattr(font, "color", None))
    if rgb and max(rgb) > 60:
        parts.append(f"글자색 #{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}")
    return " + ".join(parts) or None


def _rgb_of(color) -> Optional[tuple]:
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    try:
        return tuple(int(rgb[-6:][i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return None


def _has_color(color) -> bool:
    """기본 글자색이 아닌가. 검정·거의 검정은 표시로 보지 않는다."""
    parsed = _rgb_of(color)
    if parsed is None:
        return False
    return max(parsed) > 60          # 순검정(0,0,0)과 짙은 회색은 제외


def _is_red(color) -> bool:
    parsed = _rgb_of(color)
    if parsed is None:
        return False
    r, g, b = parsed
    return r > 140 and r > g * 1.6 and r > b * 1.6


def _merge(runs: List[dict]) -> List[dict]:
    """인접한 같은 emphasis 조각을 합쳐 노이즈를 줄인다."""
    out: List[dict] = []
    for r in runs:
        if out and out[-1]["emphasis"] == r["emphasis"]:
            out[-1]["text"] += r["text"]
        else:
            out.append(dict(r))
    return out


def cell_is_blank(c: Optional[Cell]) -> bool:
    return c is None or c.value is None or not str(c.value).strip()


def looks_numeric(v: Any) -> bool:
    if isinstance(v, bool) or isinstance(v, (dt.date, dt.datetime)):
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).strip())
        return True
    except (TypeError, ValueError):
        return False
