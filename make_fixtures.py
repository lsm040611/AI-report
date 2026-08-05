"""계약서의 원본 양식을 흉내낸 테스트 파일 생성.

    py -3.10 make_fixtures.py

fixtures/ 에 4개가 생긴다. 서로 다른 양식 4종이 같은 파이프라인을 통과하는지
확인하는 것이 목적이므로, 일부러 다음을 섞어 두었다.

  · 시리얼 날짜 (R-01)          · 문자열 점수 "3.5" (R-02)
  · 틀린 평균란 (R-03)          · 1~5 / 1~10 척도 (R-04)
  · 셀 안의 부분 서식 (R-05)    · 각주 행 (R-06)
  · 비고 열의 '청강' (R-07)     · 별도 문항정의 시트 (R-08)
  · 동료 응답 2건 (R-10)        · 제각각인 헤더 라벨 (R-12)
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")

RED = InlineFont(color="FFB3261E")
BOLD_U = InlineFont(b=True, u="single")
BOLD = InlineFont(b=True)

EXCEL_EPOCH = dt.date(1899, 12, 30)


def serial(y: int, m: int, d: int) -> int:
    """엑셀 시리얼 날짜. R-01 이 이걸 ISO 로 되돌린다."""
    return (dt.date(y, m, d) - EXCEL_EPOCH).days


def _write(wb: Workbook, name: str) -> None:
    path = os.path.join(OUT, name)
    wb.save(path)
    print(f"  · {name}")


# ══════════════════════════════════════════════════════════════
def negotiation_1() -> None:
    """1차수 — 역량 3개. 평균란이 틀려 있고(R-03), 각주 행과 청강생이 섞여 있다."""
    wb = Workbook(); ws = wb.active; ws.title = "1차수_A조"
    ws["A1"] = "과정명"; ws["B1"] = "Global Negotiation Program (GN-1)"
    ws["A2"] = "강사"; ws["B2"] = "Rachel Han"
    ws["A3"] = "날짜"; ws["B3"] = serial(2026, 5, 19)      # 시리얼 날짜 (R-01)
    ws["A4"] = "차수"; ws["B4"] = "1차수"
    ws["A6"] = "이름"
    for i, h in enumerate(["Accuracy", "Tone", "Persuasion",
                           "B-1 Strengths", "B-2 Gaps", "B-3 Next Action",
                           "평균", "비고"]):
        ws.cell(row=6, column=2 + i, value=h)

    ws["A7"] = "강지우"
    ws["B7"], ws["C7"], ws["D7"] = 4.0, "3.5", 4.5        # C7은 문자열 (R-02)
    ws["E7"] = "오프닝에서 상대 상황을 먼저 요약한 것 아주 좋았음. 신뢰를 먼저 쌓고 들어갔다."
    ws["F7"] = CellRichText([
        "가격 인상을 통보하듯 말하는 순간 있었음. ",
        TextBlock(RED, '"We will raise the price"'),
        " → ",
        TextBlock(BOLD_U, '"We may need to adjust our pricing"'),
        " 처럼 ",
        TextBlock(BOLD, "corporate softening"),
        " 필요.",
    ])
    ws["G7"] = "다음 세션까지 hedging 3종(may / could / it seems)을 의도적으로 써 볼 것."
    ws["H7"] = 4.5                                        # 틀린 평균 (실제 4.0) → R-03

    ws["A8"] = "문세린"
    ws["B8"], ws["C8"], ws["D8"] = 4.5, 4.5, 3.5
    ws["E8"] = "문법 정확도가 조에서 제일 안정적. 관사·시제 오류가 거의 없다."
    ws["F8"] = "정중한데 메시지가 각인이 안 됨. 핵심 요구를 명확히 말한 게 1번뿐."
    ws["G8"] = "결론 먼저. 첫 30초 안에 요구사항 문장이 나오는지 체크하겠음."
    ws["H8"] = 4.17

    ws["A9"] = "홍민아"
    ws["B9"], ws["C9"], ws["D9"] = 3.5, 3.0, 3.0
    ws["E9"] = "첫 참석. 관찰 위주였지만 마지막에 질문 2개 던진 것이 좋았다."
    ws["F9"] = "발화량이 적어 이번 회차는 참고 평가만 한다."
    ws["I9"] = "청강"                                      # R-07

    ws["A11"] = "※ 점수는 5점 만점 기준입니다"              # R-06 각주 행
    _write(wb, "20260519_1차수_A조.xlsx")


def negotiation_2() -> None:
    """2차수 — 역량이 5개로 늘었다. 1차수와 겹치는 3개만 성장 비교 대상(R-14)."""
    wb = Workbook(); ws = wb.active; ws.title = "2차수_A조"
    ws["A1"] = "과정명"; ws["B1"] = "Global Negotiation Program (GN-1)"
    ws["A2"] = "강사"; ws["B2"] = "Rachel Han"
    ws["A3"] = "날짜"; ws["B3"] = "2026-06-02"
    ws["A4"] = "차수"; ws["B4"] = "2차수 · A조"
    ws["A5"] = "케이스"; ws["B5"] = "물류 파트너 납기 지연 보상 협상"
    ws["A7"] = "이름"
    for i, h in enumerate(["Accuracy", "Tone", "Persuasion", "이해관계 파악", "논리 구조",
                           "B-1 Strengths", "B-2 Gaps", "B-3 Next Action"]):
        ws.cell(row=7, column=2 + i, value=h)

    ws["A8"] = "강지우"
    for col, v in zip("BCDEF", [4.0, 4.0, 4.5, 4.0, 4.5]):
        ws[f"{col}8"] = v
    ws["G8"] = CellRichText([
        "지난 시간 얘기한 ",
        TextBlock(BOLD, "hedging"),
        ' 의식적으로 쓴 게 보였음. 발전 👍 특히 "we may need to revisit the timeline" '
        "문장은 지난주 교정을 그대로 적용한 것 — 이렇게 바로 반영하는 참가자가 제일 빨리 는다.",
    ])
    ws["H8"] = CellRichText([
        "상대의 숨은 제약(성수기 인력난)을 끝까지 안 파고 보상액 논의로 직행. ",
        TextBlock(RED, "본론부터 바로 진입"),
        " 하지 말고 ",
        TextBlock(BOLD_U, '"Help me understand…"'),
        " 처럼 탐색 질문을 먼저 던질 것. 이해관계 탐색은 정보 수집이 아니라 ",
        TextBlock(BOLD, "레버리지 확보"),
        "다.",
    ])
    ws["I8"] = "본론 전에 탐색 질문 2개를 먼저 던지고 시작할 것."

    ws["A9"] = "문세린"
    for col, v in zip("BCDEF", [4.5, 4.5, 4.0, 3.5, 4.0]):
        ws[f"{col}9"] = v
    ws["G9"] = "두괄식으로 바뀌었다. 첫 40초 안에 요구사항이 나왔고 근거가 두 개 붙었다."
    ws["H9"] = "상대 반론 이후 톤이 급해졌다. 반론이 나온 뒤 한 박자 쉬는 연습이 필요하다."
    ws["I9"] = "반론 직후 3초 멈추고 되묻는 연습."

    ws["A10"] = "오태윤"
    for col, v in zip("BCDEF", [3.5, 4.0, 4.5, 4.5, 3.5]):
        ws[f"{col}10"] = v
    ws["G10"] = "상대 제약을 먼저 물어보고 들어간 유일한 참가자. 레버리지를 스스로 만들었다."
    ws["H10"] = "근거 나열이 길어 결론이 늦다. 한 문장에 이유 하나만 담을 것."
    ws["I10"] = "이유 두 개를 한 문장에 넣지 않기."
    _write(wb, "20260602_2차수_A조.xlsx")


def lecture() -> None:
    """특강: 1~10 척도, 평균란 없음, 라벨 상이 (R-03b / R-04 / R-12)"""
    wb = Workbook(); ws = wb.active; ws.title = "Presentation특강"
    ws["A1"] = "특강명"; ws["B1"] = "Executive Presentation Skills"
    ws["A2"] = "강사"; ws["B2"] = "Daniel Cho"
    ws["A3"] = "일자"; ws["B3"] = "2026-06-25"
    ws["A4"] = "장소"; ws["B4"] = "사내 교육장"
    ws["A5"] = "성명"
    for i, h in enumerate(["Structure", "Delivery", "Visuals", "Q&A",
                           "Highlights (잘한 점)", "Improvement Areas (보완점)",
                           "Homework (과제)"]):
        ws.cell(row=5, column=2 + i, value=h)
    rows = [
        ("서지호", 8, 7, 9, 6,
         "슬라이드는 오늘 참석자 중 최고. 한 장 한 메시지 원칙을 지켰다.",
         "Q&A 준비 부족 — \"확인해 보겠습니다\" 3회. 브리지 화법을 익힐 것.",
         "본인 주제의 예상 질문 5개와 답변 스크립트를 작성할 것."),
        ("한예린", 9, 8, 7, 8,
         "오프닝 후킹이 좋았다. 스토리 아크가 명확: 문제→시도→실패→전환→결과.",
         "텍스트 위주 슬라이드 2장이 흐름을 끊었다. 도식화가 필요하다.",
         "해당 2장을 다이어그램으로 리디자인해 제출할 것."),
        ("오태윤", 6, 9, 6, 7,
         "무대 장악력과 목소리는 이미 프로급. 발성이 타고난 자산이다.",
         "나열식이라 결론이 늦게 나온다. 두괄식이 필수다.",
         "발표 요지를 30초로 요약해 녹음 제출할 것."),
    ]
    for r, row in enumerate(rows, start=6):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    _write(wb, "20260625_특강.xlsx")


def survey() -> None:
    """360 진단: N응답 -> 1인 (R-09 / R-10) + 별도 문항정의 시트 (R-08)"""
    wb = Workbook(); ws = wb.active; ws.title = "응답데이터"
    ws["A1"] = "진단명"; ws["B1"] = "2026 리더십 360° 진단"
    ws["A2"] = "진단 기간"; ws["B2"] = "2026-07-01 ~ 07-10"
    ws["A4"] = "피평가자"; ws["B4"] = "관계"
    for i, q in enumerate(["Q1", "Q2", "Q3", "Q4", "Q5", "Q6"]):
        ws.cell(row=4, column=3 + i, value=q)
    ws["I4"] = "[주관식1] 이 리더의 가장 큰 강점"
    ws["J4"] = "[주관식2] 이 리더에게 바라는 변화"

    data = [
        ("한도윤", "상사", 5, 4, 4, 4, 3, 3,
         "방향 제시가 명확하고 실행이 빠릅니다.",
         "속도가 빠른 만큼 실무 의견 수렴이 생략될 때가 있습니다."),
        ("한도윤", "동료", 4, 4, 4, 3, 3, 2,
         "추진력이 좋습니다. 불확실한 국면에서 특히 그렇습니다.",
         "협업 회의 때 결론을 미리 정해놓고 오시는 느낌이 있습니다."),
        ("한도윤", "동료", 4, 4, 3, 4, 2, 3,
         "판단이 빠르고 지시가 구체적입니다.",
         "결정 전에 실무 의견을 듣는 단계가 한 번 있으면 좋겠습니다."),
        ("한도윤", "구성원", 4, 4, 4, 4, 3, 2,
         "디렉션이 구체적이라 헤맬 일이 없습니다.",
         "저희 팀에서는 작년 워크숍 때 제가 낸 의견이 그냥 넘어가서 좀 그랬어요ㅠ"),
        ("한도윤", "구성원", 5, 5, 4, 3, 2, 2,
         "교육 기회를 잘 챙겨주십니다.",
         "1on1이 자주 밀려요ㅠ 세 번 연속 밀리니까 좀 그랬습니다."),
        ("한도윤", "구성원", 4, 4, 3, 4, 3, 3,
         "전문성이 뛰어나고 실행이 지체되지 않습니다.",
         "보고 중간에 말을 끊지 말아주세요. 끝까지 듣고 판단해 주시면 좋겠습니다."),
    ]
    for r, row in enumerate(data, start=5):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)

    # 문항 정의는 별도 시트에 있다 (R-08)
    d = wb.create_sheet("문항정의")
    d["A1"] = "문항번호"; d["B1"] = "역량"; d["C1"] = "정의문"
    defs = [
        ("Q1", "비전 제시", "팀이 나아갈 방향과 목표를 명확하게 제시한다"),
        ("Q2", "성장 지원", "구성원 개개인의 성장과 경력 개발을 적극 지원한다"),
        ("Q3", "공정성", "평가·보상·업무 배분에서 일관되고 공정하다"),
        ("Q4", "피드백 제공", "구체적이고 도움이 되는 피드백을 적시에 제공한다"),
        ("Q5", "위임과 권한부여", "구성원을 신뢰하고 실질적인 권한을 위임한다"),
        ("Q6", "경청·소통", "구성원의 의견을 끝까지 듣고 열린 자세로 소통한다"),
    ]
    for r, row in enumerate(defs, start=2):
        for c, v in enumerate(row, start=1):
            d.cell(row=r, column=c, value=v)

    _write(wb, "20260710_360진단.xlsx")


if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    print(f"fixtures 생성 → {OUT}")
    negotiation_1()
    negotiation_2()
    lecture()
    survey()
    print("완료. 서버를 띄운 뒤 http://127.0.0.1:8000 에서 업로드해 보세요.")
