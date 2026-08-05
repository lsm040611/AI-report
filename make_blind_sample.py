"""블라인드 테스트용 샘플 평가지를 만든다 — 배포 샘플과 **일부러 다르게** 생겼다.

최종 발표에서는 처음 보는 양식이 들어온다. 그래서 배포 샘플과 같은 모양으로만
맞춰 두면 그날 무너진다. 이 파일은 다음을 전부 바꿔서 만든다.

    과정·강사·참가자        전부 새 인물
    척도                   1~5 → **1~7점**, 각주 문구도 다르게
    역량명                 처음 보는 이름 4개 (배포 샘플은 3개)
    열 구성                사번·이메일·소속 열 추가, 순서도 다름
    서술 칸 이름           B-1/B-2/B-3 → 한국어 제목
    강조 색                빨강/굵게+밑줄 → **보라 / 초록 굵게 / 밑줄만**

그리고 실제 업무 데이터가 그렇듯 **이상한 곳을 일부러 심어 둔다.** 평균 네 칸이
각각 다른 상태(정상·오류·수식만·빈칸)이고, 점수는 문자열과 숫자가 섞여 있으며,
한 사람은 코멘트가 통째로 영어다. 이름 표기도 세 가지가 섞여 있다.

    py -3.10 make_blind_sample.py [내보낼 폴더]
"""
from __future__ import annotations

import os
import sys

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill

# 배포 샘플의 리포트를 받는 사람. 더미 데이터이고, 생성된 리포트를 실제로
# 받아 보기 위한 주소라 네 명 모두 같은 값을 쓴다.
TESTER_EMAIL = "lsm040611@naver.com"

OUT_DEFAULT = (r"C:\Users\user\OneDrive\문서\카카오톡 받은 파일"
               r"\2팀_샘플데이터_배포패키지_3\SUNIC_2팀_샘플데이터")
FILENAME = "20260805_3차수_C조.xlsx"

# ── 강조 서식 — 배포 샘플과 뜻이 같지만 색이 다르다 ────────────────────────
FIX_BAD = InlineFont(color="FF7B2D8E")                 # 고칠 표현  (배포본은 빨강)
FIX_GOOD = InlineFont(b=True, color="FF1B7F4B")        # 권장 표현  (배포본은 굵게+밑줄)
KEY = InlineFont(u="single")                           # 핵심 개념  (배포본은 굵게)

# ── 시트 서식 ──────────────────────────────────────────────────────────────
TITLE_FILL = PatternFill("solid", fgColor="FF3C2A21")
LABEL_FILL = PatternFill("solid", fgColor="FFF4EDE4")
HEAD_FILL = PatternFill("solid", fgColor="FF6B4F3A")
WHITE_BOLD = Font(bold=True, color="FFFFFFFF")
BOLD = Font(bold=True)
MUTED = Font(color="FF6B7280", size=9)

AREAS = [
    ("Framing", "문제를 어떻게 규정했는가"),
    ("Evidence", "근거의 적절성과 출처"),
    ("Visual", "도표 선택과 가독성"),
    ("Delivery", "전달 속도와 시선 처리"),
]

NARRATIVE = ["잘한 점", "아쉬운 점", "다음까지 할 것"]


def _people(ws) -> None:
    """네 사람. 이름 표기·평균 상태·점수 자료형을 일부러 다르게 둔다."""

    # ── 윤태오 : 정상. 이름만, 평균은 하드코딩된 정확한 값 ──────────────
    ws["A7"], ws["B7"], ws["C7"], ws["D7"] = "윤태오", "E20811", TESTER_EMAIL, "전략기획팀"
    ws["E7"], ws["F7"], ws["G7"], ws["H7"] = 5.5, 6, 5, 6.5      # 숫자형
    ws["I7"] = 5.8                                                # 평균: 정상
    ws["J7"] = CellRichText([
        "첫 3분에 ", TextBlock(KEY, "문제 정의"), "를 끝낸 구성이 좋았습니다. ",
        "청중이 무엇을 결정해야 하는지 먼저 알고 들었습니다.",
    ])
    ws["K7"] = CellRichText([
        "근거를 말할 때 ", TextBlock(FIX_BAD, "\"데이터를 보면 아시겠지만\""),
        " 처럼 넘어간 구간이 두 번 있었습니다. ",
        TextBlock(FIX_GOOD, "\"이 표에서 보실 부분은 세 번째 행입니다\""),
        " 처럼 짚어 주면 청중이 따라옵니다.",
    ])
    ws["L7"] = CellRichText([
        "다음 발표까지 ", TextBlock(KEY, "출처 한 줄"), "을 모든 도표에 답니다. ",
        "숫자를 의심받는 순간 발표가 멈춥니다.",
    ])

    # ── 배시현 : 평균이 하드코딩인데 **틀렸다** (엔진이 잡아야 한다) ────
    ws["A8"], ws["B8"], ws["C8"], ws["D8"] = "배시현 (Sarah)", "E20744", TESTER_EMAIL, "마케팅팀"
    ws["E8"], ws["F8"], ws["G8"], ws["H8"] = "4", "4.5", "3.5", "4"   # 문자열형
    ws["I8"] = 5.2                                                    # 실제 평균은 4.0
    ws["J8"] = CellRichText([
        "도표를 직접 다시 그려 온 성의가 보였습니다. ",
        TextBlock(KEY, "축 단위 통일"), "만으로 앞 장과 비교가 쉬워졌습니다.",
    ])
    ws["K8"] = CellRichText([
        "한 장에 담은 정보가 많습니다. ", TextBlock(FIX_BAD, "막대 + 꺾은선 + 표"),
        " 를 한 화면에 두면 어디를 보라는 것인지 알기 어렵습니다. ",
        TextBlock(FIX_GOOD, "한 장 한 메시지"), " 원칙을 권합니다.\n",
        "질문을 받았을 때 ", TextBlock(FIX_BAD, "\"그건 제가 확인해서 말씀드릴게요\""),
        " 로 넘긴 것이 세 번 있었습니다.",
    ])
    ws["L8"] = CellRichText([
        "예상 질문 다섯 개를 미리 적고 답을 준비합니다. ",
        TextBlock(KEY, "모르면 모른다고 말하기"), "도 답의 하나입니다.",
    ])

    # ── 조은결 : 평균이 수식 (계산값이 저장돼 있지 않다) ────────────────
    ws["A9"], ws["B9"], ws["C9"], ws["D9"] = "조은결/Eugene", "E20902", TESTER_EMAIL, "재무팀"
    ws["E9"], ws["F9"], ws["G9"], ws["H9"] = 6, "6.5", 6, 5.5        # 자료형 혼재
    ws["I9"] = "=ROUND(AVERAGE(E9:H9),1)"                             # 평균: 수식만
    ws["J9"] = CellRichText([
        "숫자를 다루는 손이 빠릅니다. 질문이 들어와도 ",
        TextBlock(KEY, "원자료로 되돌아가는 습관"), "이 있어 답이 흔들리지 않았습니다.",
    ])
    ws["K9"] = CellRichText([
        "설명이 회계 용어에 기대 있습니다. ", TextBlock(FIX_BAD, "\"영업레버리지가 높아서\""),
        " 대신 ", TextBlock(FIX_GOOD, "\"매출이 조금만 늘어도 이익이 크게 늘어서\""),
        " 처럼 풀어 주면 비재무 청중도 따라옵니다.",
    ])
    ws["L9"] = CellRichText([
        "이번 자료에서 용어 세 개를 골라 ", TextBlock(KEY, "한 문장 설명"), "을 붙여 옵니다.",
    ])

    # ── 남지후 : 평균 칸이 비어 있고, 점수 한 칸도 결측. 코멘트는 영어 ──
    ws["A10"], ws["B10"], ws["C10"], ws["D10"] = "남지후", "E20655", TESTER_EMAIL, "영업1팀"
    ws["E10"], ws["F10"], ws["G10"] = 4.5, 5, None                    # G10 결측
    ws["H10"] = 5
    # I10 은 비워 둔다 — 평균란 자체가 없는 경우
    ws["J10"] = CellRichText([
        "Opened with a concrete customer story. ",
        TextBlock(KEY, "story-first framing"),
        " made the numbers easier to accept.",
    ])
    ws["K10"] = CellRichText([
        "Tends to read the slide aloud. ",
        TextBlock(FIX_BAD, "reading every bullet"), " → ",
        TextBlock(FIX_GOOD, "say the takeaway, then point at the evidence"), ".\n",
        "Also spoke noticeably faster in the last two minutes.",
    ])
    ws["L10"] = CellRichText([
        "Rehearse the closing once out loud, timed. ",
        TextBlock(KEY, "one message per slide"), " applies to the summary slide too.",
    ])


def build(out_dir: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "3차수_C조"

    # ── 머리말 ──────────────────────────────────────────────────────────
    ws["A1"] = "Data Storytelling Intensive — 세션 피드백"
    ws["A1"].font, ws["A1"].fill = WHITE_BOLD, TITLE_FILL
    ws.merge_cells("A1:L1")

    meta = [
        ("A2", "과정명", "B2", "Data Storytelling Intensive (DS-2)"),
        ("D2", "강사", "E2", "서다인"),
        ("A3", "날짜", "B3", "2026-08-05"),
        ("D3", "차수", "E3", "3차수"),
        ("G3", "조", "H3", "C조"),
        ("A4", "세션 주제", "B4", "분기 실적 보고를 임원에게 5분 안에 전달하기"),
    ]
    for lc, label, vc, value in meta:
        ws[lc], ws[vc] = label, value
        ws[lc].font, ws[lc].fill = BOLD, LABEL_FILL

    # ── 표 머리 ─────────────────────────────────────────────────────────
    headers = (["이름", "사번", "이메일", "소속"]
               + [f"{en}\n{ko}" for en, ko in AREAS]
               + ["평균"] + NARRATIVE)
    for idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=idx, value=text)
        cell.font, cell.fill = WHITE_BOLD, HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    _people(ws)

    # ── 각주 : 척도가 표 밖에만 적혀 있다 ───────────────────────────────
    ws["A12"] = "※ 평가 척도는 7점 만점이며 0.5점 단위로 기재합니다. 미실시 항목은 공란."
    ws["A12"].font = MUTED

    for col, width in zip("ABCDEFGHIJKL",
                          [16, 10, 24, 12, 10, 10, 10, 10, 8, 44, 52, 44]):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[6].height = 32

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, FILENAME)
    wb.save(path)
    return path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else OUT_DEFAULT
    print("생성:", build(out))
