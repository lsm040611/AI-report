"""까다로운 양식 회귀 테스트.

    py -3.10 test_edge_cases.py

실제로 터졌던 케이스를 고정해 둔다. 특히 **셀 부분 서식이 머리부에 들어간 파일**은
카드가 DB에 저장되는 순간 `TypeError: Object of type TextBlock is not JSON
serializable` 로 500을 냈다. 서식은 서술 열에만 있을 거라고 가정한 탓이었다.
"""
from __future__ import annotations

import datetime as dt
import json
import os
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
os.chdir(HERE)
os.environ["HR_DB_URL"] = "sqlite:///./edge.db"
os.environ["HR_AUTO_APPROVE"] = "1"
if os.path.exists("edge.db"):
    os.remove("edge.db")

from openpyxl import Workbook                                    # noqa: E402
from openpyxl.cell.rich_text import CellRichText, TextBlock      # noqa: E402
from openpyxl.cell.text import InlineFont                        # noqa: E402

from pipeline import run_pipeline                                # noqa: E402

BOLD = InlineFont(b=True)
RED = InlineFont(color="FFB3261E")
TMP = tempfile.mkdtemp(prefix="hr_edge_")

_failures = []


def check(label, cond, detail=""):
    print(("  OK  " if cond else "  !!  ") + label + (f"  {detail}" if detail else ""))
    if not cond:
        _failures.append(label)


def run(name: str, build) -> dict:
    """워크북을 만들고 파이프라인 + JSON 직렬화까지 통과하는지 본다."""
    wb = Workbook()
    build(wb)
    path = os.path.join(TMP, name)
    wb.save(path)
    result = run_pipeline(path, auto_approve=True)
    json.dumps(result["cards"], ensure_ascii=False)     # 저장 가능해야 한다
    json.dumps(result["handoffs"], ensure_ascii=False)
    return result


# ══════════════════════════════════════════════════════════════
def case_rich_meta():
    """머리부 셀에 부분 서식이 들어간 파일 — 실제로 500을 냈던 케이스."""
    def build(wb):
        ws = wb.active
        ws.title = "평가"
        ws["A1"] = "과정명"
        ws["B1"] = CellRichText(["Global ", TextBlock(BOLD, "Negotiation"), " Program"])
        ws["A2"] = "차수"
        ws["B2"] = CellRichText([TextBlock(RED, "2차수"), " · A조"])
        ws["A3"] = "일자"
        ws["B3"] = dt.datetime(2026, 6, 2, 14, 30)      # datetime 도 그대로 들어온다
        ws["A5"] = "이름"
        for i, h in enumerate(["Tone", "Accuracy", "강점", "보완점"]):
            ws.cell(row=5, column=2 + i, value=h)
        ws["A6"], ws["B6"], ws["C6"] = "김하늘", 4.0, 3.5
        ws["D6"] = "상대 상황을 먼저 요약하고 들어간 점이 좋았습니다."
        ws["E6"] = "결론이 늦게 나옵니다. 두괄식으로 말할 것."
        ws["A7"], ws["B7"], ws["C7"] = "박서준", 3.5, 4.0
        ws["D7"] = "근거 제시가 구체적이고 숫자를 함께 씁니다."
        ws["E7"] = "상대 반론 이후 톤이 급해집니다."

    r = run("rich_meta.xlsx", build)
    check("카드 2장 생성", r["summary"]["cards"] == 2, r["summary"]["cards"])
    ctx = r["cards"][0]["context"]
    check("서식 섞인 과정명이 평문으로", ctx.get("과정명") == "Global Negotiation Program",
          repr(ctx.get("과정명")))
    check("서식 섞인 차수가 평문으로", ctx.get("차수") == "2차수 · A조", repr(ctx.get("차수")))
    check("datetime 이 ISO 문자열로", isinstance(ctx.get("일자"), str), repr(ctx.get("일자")))
    check("누적교육으로 판정", r["cards"][0]["source_type"]["type"] == "누적교육")


def case_no_scores():
    """점수가 전부 비어 있고 서술만 있는 사람."""
    def build(wb):
        ws = wb.active
        ws["A1"] = "특강명"; ws["B1"] = "테스트 특강"
        ws["A3"] = "이름"
        ws["B3"], ws["C3"] = "Delivery", "총평"
        ws["A4"], ws["B4"] = "이도현", 7
        ws["C4"] = "발성이 안정적이고 시선 처리가 좋았습니다."
        ws["A5"] = "최유진"                       # 점수 없음, 서술만
        ws["C5"] = "관찰 위주였지만 마지막 질문이 좋았습니다."

    r = run("no_scores.xlsx", build)
    names = [c["person"]["name"] for c in r["cards"]]
    check("점수 없는 사람도 카드가 됨(R-06: 서술이 있으면 사람)", "최유진" in names, names)
    blank = [c for c in r["cards"] if c["person"]["name"] == "최유진"][0]
    check("평균은 계산하지 않음", blank["score_summary"].get("average") is None,
          blank["score_summary"])


def case_cover_sheet_first():
    """첫 시트가 표지·안내문이고 데이터는 두 번째 시트에 있는 경우."""
    def build(wb):
        cover = wb.active
        cover.title = "표지"
        cover["A1"] = "2026년 상반기 교육 결과 보고"
        cover["A3"] = "본 문서는 대외비입니다."
        ws = wb.create_sheet("결과")
        ws["A1"] = "과정명"; ws["B1"] = "리더십 기초"
        ws["A3"] = "성명"; ws["B3"] = "이해도"; ws["C3"] = "잘한 점"
        ws["A4"], ws["B4"], ws["C4"] = "정민서", 4, "질문이 핵심을 짚습니다."
        ws["A5"], ws["B5"], ws["C5"] = "한지원", 5, "사례 연결이 정확합니다."

    r = run("cover_first.xlsx", build)
    check("표지 시트를 건너뛰고 데이터 시트를 찾음", r["summary"]["cards"] == 2,
          r["summary"])
    check("사용한 시트가 '결과'", r["summary"]["sheets"] == ["결과"], r["summary"]["sheets"])


def case_merged_and_gaps():
    """병합 셀(뒤 칸이 None)과 빈 행이 섞인 경우."""
    def build(wb):
        ws = wb.active
        ws["A1"] = "과정명"; ws["B1"] = "협상 실무"
        ws.merge_cells("B1:D1")
        ws["A3"] = "이름"; ws["B3"] = "Tone"; ws["C3"] = "총평"
        ws["A4"], ws["B4"], ws["C4"] = "오세훈", 4.0, "톤이 일관됩니다."
        # 5행은 통째로 비움
        ws["A6"], ws["B6"], ws["C6"] = "윤아름", 3.5, "속도 조절이 필요합니다."

    r = run("merged.xlsx", build)
    check("병합·빈 행이 있어도 카드 2장", r["summary"]["cards"] == 2, r["summary"]["cards"])


def case_no_name_column():
    """이름 열이 없는 요약표 — 죽지 않되, 추측했다는 사실이 담당자에게 올라가야 한다."""
    def build(wb):
        ws = wb.active
        ws["A1"] = "항목"; ws["B1"] = "값"
        ws["A2"], ws["B2"] = "총 참석", 24
        ws["A3"], ws["B3"] = "만족도", 4.6

    r = run("no_name.xlsx", build)
    check("예외 없이 끝남", True)
    guessed = [w for w in r["warnings"] if "추측" in w or "찾지 못" in w]
    check("이름 열을 추측했다는 경고가 올라옴", bool(guessed), r["warnings"])


def case_growth_mapping_drift():
    """1차수부터 있던 역량이 '새로 추가'로 잘못 표시되던 문제.

    표준 역량명(R-18)은 담당자 승인 여부에 따라 업로드 시점마다 달라진다.
    1차수를 올릴 땐 매핑이 없어 canonical 이 비어 있고, 2차수를 올릴 땐
    승인된 표준명이 붙는다. 이름 하나만으로 대조하면 같은 역량이 새 역량이 된다.
    """
    from pipeline.rules.report import r14_growth

    prev = {                                   # 1차수 — 매핑 승인 전
        "scores": [
            {"area_name": "Persuasion", "canonical_area": None, "score": 4.0,
             "scale": {"min": 1, "max": 5, "step": 0.5}},
            {"area_name": "이해관계 파악", "canonical_area": None, "score": 3.5,
             "scale": {"min": 1, "max": 5, "step": 0.5}},
        ],
        "score_summary": {"average": 3.75},
    }
    cur = {                                    # 2차수 — 매핑 승인 후 + 띄어쓰기 차이
        "scores": [
            {"area_name": "Persuasion", "canonical_area": "설득력", "score": 4.5,
             "scale": {"min": 1, "max": 5, "step": 0.5}},
            {"area_name": "이해관계파악", "canonical_area": "이해관계파악", "score": 4.0,
             "scale": {"min": 1, "max": 5, "step": 0.5}},
            {"area_name": "논리 구조", "canonical_area": "논리구조", "score": 4.0,
             "scale": {"min": 1, "max": 5, "step": 0.5}},
        ],
        "score_summary": {"average": 4.17},
    }

    g = r14_growth(cur, prev)
    labels = {d["label"]: d["delta"] for d in g["deltas"]}
    new = [n["label"] for n in g["new_areas"]]

    check("표준명이 바뀌어도 같은 역량으로 대조", "Persuasion" in labels, labels)
    check("띄어쓰기가 달라도 같은 역량으로 대조", "이해관계파악" in labels, labels)
    check("Persuasion 증감 +0.5", labels.get("Persuasion") == 0.5, labels)
    check("진짜 신규 역량만 신규로", new == ["논리 구조"], new)
    check("구성이 다르므로 평균 비교 미표시", g["average_comparable"] is False)


def case_action_items_from_emphasis():
    """실천 항목이 강조 표시(색·굵기·밑줄)에서 나오고 말투가 일관되는지."""
    from generation.worker import _mock_r17

    payload = {
        "form": "연계형",
        "evidence": [{"role": "gap", "quote": "가격을 통보하듯 말한 순간이 있었습니다."}],
        "pairs": [{"issue": "We will raise the price",
                   "fix": "We may need to adjust our pricing", "role": "gap"}],
        "emphasis": [
            {"kind": "issue", "text": "We will raise the price", "role": "gap"},
            {"kind": "fix", "text": "We may need to adjust our pricing", "role": "gap"},
            {"kind": "key", "text": "hedging", "role": "gap"},
        ],
    }
    out = _mock_r17(payload)
    lines = [ln for ln in out["text"].splitlines() if ln.strip()]

    check("항목이 만들어짐", len(lines) >= 2, lines)
    check("교정 쌍이 첫 항목", "대신" in lines[0], lines[0])
    check("강조 표현이 원문 그대로 실림",
          "We may need to adjust our pricing" in lines[0], lines[0])
    check("말투가 '~합니다' 로 통일", all(ln.rstrip().endswith("니다.") for ln in lines),
          lines)
    check("영문 받침 조사 처리 (hedging을)",
          any("hedging</b>을" in ln for ln in lines), lines)
    check("근거가 항목 수만큼", len(out["evidence"]) == len(lines))


def case_gap_comment_rewrite():
    """'함께 살펴보면 좋을 점'에 덧붙는 코멘트가 원문 복사가 아닌지."""
    from generation.worker import _mock_gap

    original = "상대의 숨은 제약을 끝까지 안 파고 보상액 논의로 직행."
    out = _mock_gap({
        "gap_text": original,
        "weakest": {"name": "이해관계 파악", "score": 3.5, "max": 5},
        "pairs": [{"issue": "본론부터 바로 진입",
                   "fix": "Help me understand…", "role": "gap"}],
        "emphasis": [{"kind": "key", "text": "레버리지 확보", "role": "gap"}],
    })
    text = out["text"]

    check("문단이 만들어짐", len(text) > 40, text[:60])
    check("원문을 그대로 옮기지 않음", original not in text)
    check("강조 표현은 원문 그대로 인용", "Help me understand…" in text)
    check("모든 문장이 '~니다.' 로 끝남",
          all(s.strip().endswith("니다") for s in text.rstrip(".").split(". ")), text)
    check("근거가 붙음(R-16)", bool(out["evidence"]), out["evidence"])


def case_tone_normalisation():
    """강사 메모체·요청형이 존댓말로 통일되는지."""
    from tone import polish as _polish

    cases = {
        "상대 반론 이후 톤이 급해졌다": "상대 반론 이후 톤이 급해졌습니다.",
        "한 박자 쉬는 연습이 필요하다": "한 박자 쉬는 연습이 필요합니다.",
        "두괄식이 필수다": "두괄식이 필수입니다.",
        "이유 두 개를 한 문장에 넣지 않기": "이유 두 개를 한 문장에 넣지 않습니다.",
        "예상 질문 5개를 작성할 것": "예상 질문 5개를 작성합니다.",
        "보고 중간에 말을 끊지 말아주세요": "보고 중간에 말을 끊지 않습니다.",
        "끝까지 듣고 판단해 주시면 좋겠습니다": "끝까지 듣고 판단합니다.",
        "반론 직후 3초 멈추고 되묻는 연습": "반론 직후 3초 멈추고 되묻는 연습을 합니다.",
        "오류가 거의 없습니다": "오류가 거의 없습니다.",
    }
    for src, want in cases.items():
        got = _polish(src)
        check(f"{src[:18]}…", got == want, f"→ {got}")


def case_two_line_header_and_alias():
    """머리글이 두 줄이고 이름 칸에 별칭이 붙은 실제 양식.

    '이름 (Aiden)' 과 'Accuracy\\nGrammar & Vocabulary' 형태. 둘째 줄이 회차마다
    한국어→영어로 바뀌면, 줄바꿈째로 역량명을 잡을 때 같은 역량이 다른 역량이 되어
    성장 비교가 통째로 어긋난다.
    """
    from pipeline.rules.report import r14_growth

    def build(wb, second_line, extra):
        ws = wb.active
        ws["A1"] = "과정명"; ws["B1"] = "Global Negotiation Program (GN-1)"
        ws["A2"] = "차수"; ws["B2"] = extra
        ws["A3"] = "날짜"; ws["B3"] = "2026-05-21" if extra == "1차수" else "2026-06-04"
        ws["A5"] = "이름"
        cols = ["Accuracy\n" + second_line[0], "Tone\n" + second_line[1], "총평"]
        for i, h in enumerate(cols):
            ws.cell(row=5, column=2 + i, value=h)
        ws["A6"] = "서준혁 (Aiden)"
        ws["B6"], ws["C6"] = (3.5, 4.0) if extra == "1차수" else (4.0, 4.5)
        ws["D6"] = "문장이 짧아지고 요지가 먼저 나옵니다."
        ws["A7"] = "한지원 (Grace)"
        ws["B7"], ws["C7"] = (4.0, 3.5) if extra == "1차수" else (4.0, 4.0)
        ws["D7"] = "근거 제시가 구체적입니다."

    r1 = run("two_line_1.xlsx", lambda wb: build(wb, ("문법·어휘 정확성", "외교적 어조"), "1차수"))
    r2 = run("two_line_2.xlsx", lambda wb: build(wb, ("Grammar & Vocabulary", "Diplomatic tone"), "2차수"))

    p = r2["cards"][0]["person"]
    check("이름과 별칭 분리", (p["name"], p["alias"]) == ("서준혁", "Aiden"), p)

    s = r2["cards"][0]["scores"][0]
    check("역량명은 첫 줄만", s["area_name"] == "Accuracy", s["area_name"])
    check("둘째 줄은 정의문으로", s["definition"] == "Grammar & Vocabulary", s.get("definition"))
    check("표준 역량 매핑됨", s["canonical_area"] == "정확성", s.get("canonical_area"))

    prev = {c["person"]["name"]: c for c in r1["cards"]}
    g = r14_growth(r2["cards"][0], prev["서준혁"])
    labels = {d["label"]: d["delta"] for d in g["deltas"]}
    check("둘째 줄이 바뀌어도 같은 역량으로 대조",
          set(labels) == {"Accuracy", "Tone"}, labels)
    check("신규 역량 없음", g["new_areas"] == [], g["new_areas"])


def case_english_comment_translation():
    """영어 강사 코멘트는 번역이 붙으면 한국어가 본문이 된다."""
    from render.adapter import Sentences, _fill_body

    nar = {"language": "en", "translation_ko": "마무리를 서두르는 장면이 있었습니다.",
           "runs": [{"text": "You still rush the closing.", "emphasis": None}]}
    sid = Sentences({"provenance": {"file": "평가표.xlsx", "sheet": "3차", "row": 21}})
    section = {}
    _fill_body(section, nar, nar["runs"], sid)
    check("번역이 본문", "마무리를 서두르는" in (section.get("html") or ""), section.get("html"))
    check("영어 원문은 아래에 남음",
          "rush the closing" in (section.get("notes") or [{}])[0].get("html", ""),
          section.get("notes"))
    check("번역문에 문장 id 가 붙음", section.get("sid") == "s1", section.get("sid"))
    check("근거는 영어 원문", "rush the closing" in sid.items[0]["sourceText"],
          sid.items[0])
    check("출처 표기", sid.items[0]["sourceRef"] == "평가표.xlsx › 3차 › 21행",
          sid.items[0]["sourceRef"])

    # 번역이 아직 없으면 원문을 그대로 싣는다
    section2 = {}
    _fill_body(section2, {"language": "en", "runs": nar["runs"]}, nar["runs"], sid)
    check("번역 없으면 원문 유지", "runs" in section2 and "html" not in section2, section2)
    check("사람이 쓴 원문에는 id 를 붙이지 않음", "sid" not in section2, section2)


def case_emphasis_reinterpreted():
    """강사가 좋은 점·나쁜 점을 가리지 않고 전부 굵게 칠한 경우.

    서식을 그대로 의미로 옮기면 '고칠 표현'과 '권장 표현'이 뒤바뀐다.
    문장을 읽고 판정한 결과가 서식을 덮어써야 한다.
    """
    from generation.runner import _apply_emphasis
    from pipeline.rules.structural import retag_runs

    text = ('가격을 통보하듯 말한 순간이 있었습니다. "We will raise the price" 대신 '
            '"We may need to adjust our pricing" 를 쓰십시오.')
    card = {"narratives": [{
        "original_label": "B-2 Gaps",
        "role": "gap",
        # 강사는 둘 다 그냥 굵게만 칠했다 → 서식상 둘 다 key_concept
        "runs": [
            {"text": '가격을 통보하듯 말한 순간이 있었습니다. ', "emphasis": None},
            {"text": '"We will raise the price"', "emphasis": "key_concept"},
            {"text": ' 대신 ', "emphasis": None},
            {"text": '"We may need to adjust our pricing"', "emphasis": "key_concept"},
            {"text": ' 를 쓰십시오.', "emphasis": None},
        ],
    }], "flags": []}

    result = {"text": "교정 쌍입니다.", "spans": [
        {"quote": '"We will raise the price"', "kind": "issue", "why": "고치라고 지적"},
        {"quote": '"We may need to adjust our pricing"', "kind": "fix",
         "why": "대신 쓰라고 제시"},
    ]}
    _apply_emphasis(card, {"label": "B-2 Gaps", "text": text}, result)

    got = {r["text"]: r.get("emphasis") for r in card["narratives"][0]["runs"]}
    check("고칠 표현으로 재분류",
          got.get('"We will raise the price"') == "issue_expression", got)
    check("권장 표현으로 재분류",
          got.get('"We may need to adjust our pricing"') == "corrected_expression", got)
    check("판정 출처가 ai", card["narratives"][0].get("emphasis_source") == "ai")
    check("바뀐 구간을 기록", len(card["narratives"][0].get("emphasis_changed") or []) == 2,
          card["narratives"][0].get("emphasis_changed"))
    check("담당자용 플래그", any(f["code"] == "emphasis_reinterpreted"
                          for f in card["flags"]), card["flags"])
    check("원문 글자는 그대로",
          "".join(r["text"] for r in card["narratives"][0]["runs"]) == text)

    # 겹치는 구간과 원문에 없는 구간은 버린다
    runs = retag_runs("abc def ghi", [
        {"quote": "abc def", "kind": "issue", "why": ""},
        {"quote": "def", "kind": "fix", "why": ""},          # 겹침 → 버림
        {"quote": "zzz", "kind": "key", "why": ""},          # 없음 → 버림
    ])
    check("겹침·미존재 구간 처리",
          [(r["text"], r["emphasis"]) for r in runs] ==
          [("abc def", "issue_expression"), (" ghi", None)], runs)


def case_blind_new_format():
    """처음 보는 양식 — 최종 발표 블라인드 테스트를 흉내낸 것.

    지금까지 쓰던 양식과 일부러 전부 다르게 만든다: 척도 1~7, 머리부 3열 배치,
    평균이 앞쪽, 처음 보는 역량명, 이름 칸에 사번.
    """
    def build(wb):
        ws = wb.active
        ws.title = "C조 평가표"
        ws["A1"] = "프로그램"; ws["B1"] = "Customer Excellence Lab"
        ws["D1"] = "운영"; ws["E1"] = "CX운영팀"
        ws["A2"] = "퍼실리테이터"; ws["B2"] = "정하윤"
        ws["A3"] = "실시일"; ws["B3"] = "2026-08-12"
        ws["D3"] = "회차"; ws["E3"] = "3회차"
        for i, h in enumerate(["참가자", "종합 평균", "공감 표현\nEmpathy",
                               "문제 정의\nProblem framing", "대안 제시\nOption giving",
                               "잘한 점", "다음에 더 좋아질 점", "메모"]):
            ws.cell(row=5, column=1 + i, value=h)

        ws["A6"] = "임채원 (E20417)"
        ws["B6"] = 6.0
        ws["C6"], ws["D6"], ws["E6"] = 6.0, 5.5, 6.5
        ws["F6"] = "고객이 말을 끝내기 전에 끼어들지 않았습니다."
        ws["G6"] = "해결책을 먼저 던지는 장면이 있었습니다."

        ws["A7"] = "노시우 (E20418)"
        ws["B7"] = 5.2
        ws["C7"], ws["D7"], ws["E7"] = 5.0, 5.5, 5.0
        ws["F7"] = "질문이 구체적입니다."
        ws["G7"] = "말이 빨라 고객이 되묻는 일이 있었습니다."

        ws["A8"] = "백지훈 (E20419)"
        ws["B8"] = 6.2
        ws["C8"], ws["D8"] = 6.5, 6.0          # 대안 제시 결측(조퇴)
        ws["F8"] = "대안을 두 개씩 제시했습니다."
        ws["G8"] = "마무리 확인이 생략된 통화가 있었습니다."
        ws["H8"] = "조퇴"

        ws["A10"] = "※ 7점 만점, 0.5 단위로 평가합니다"   # 각주 행 + 척도 힌트

    r = run("blind_new_format.xlsx", build)

    check("각주 행을 빼고 3장", r["summary"]["cards"] == 3, r["summary"]["cards"])
    card = r["cards"][0]
    p = card["person"]
    check("사번은 신원 키로 (제목에 붙지 않음)",
          (p["name"], p["alias"], p["person_id"]) == ("임채원", None, "E20417"), p)

    scale = card["score_summary"]["scale"]
    check("'7점 만점' 문구로 척도 인식", (scale["min"], scale["max"]) == (1, 7), scale)

    s = card["scores"][0]
    check("역량명은 첫 줄만", s["area_name"] == "공감 표현", s["area_name"])
    check("둘째 줄은 정의문", s["definition"] == "Empathy", s.get("definition"))
    check("평균 열은 역량에서 제외", len(card["scores"]) == 3,
          [x["area_name"] for x in card["scores"]])
    check("원본 평균을 따로 보관", card["score_summary"]["original_average"] == 6.0,
          card["score_summary"])

    missing = [c for c in r["cards"] if c["person"]["name"] == "백지훈"][0]
    gap = [x for x in missing["scores"] if x["area_name"] == "대안 제시"][0]
    check("결측은 0이 아니라 미평가", gap["score"] is None, gap)


def case_multi_file_upload():
    """여러 파일을 한 번에 — 순서를 거꾸로 올려도 성장 비교가 붙어야 한다.

    파일을 하나씩 처리하며 리포트까지 만들면, 2차수를 먼저 올린 경우 비교할
    1차수가 아직 없어 성장 섹션이 빠진다. 카드를 모두 만든 뒤에 리포트를
    만들어야 순서와 무관해진다.
    """
    import datetime as _dt
    from fastapi.testclient import TestClient
    from main import app

    def build(wb, when, label, tone_scores):
        ws = wb.active
        ws["A1"] = "과정명"; ws["B1"] = "Multi Upload Program"
        ws["A2"] = "차수"; ws["B2"] = label
        ws["A3"] = "날짜"; ws["B3"] = when
        ws["A5"] = "이름"
        for i, h in enumerate(["Tone", "Accuracy", "잘한 점"]):
            ws.cell(row=5, column=2 + i, value=h)
        for r, (name, tone, acc) in enumerate(tone_scores, start=6):
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=tone)
            ws.cell(row=r, column=3, value=acc)
            ws.cell(row=r, column=4, value="정리가 명확했습니다.")

    first = os.path.join(TMP, "multi_1차수.xlsx")
    second = os.path.join(TMP, "multi_2차수.xlsx")
    for path, when, label, rows in (
            (first, "2026-03-02", "1차수", [("김유나", 3.5, 4.0), ("박建", 4.0, 3.5)]),
            (second, "2026-03-16", "2차수", [("김유나", 4.0, 4.0), ("박建", 4.0, 4.5)])):
        wb = Workbook()
        build(wb, when, label, rows)
        wb.save(path)

    client = TestClient(app)
    # 일부러 2차수를 먼저 올린다
    payload = [("file", (os.path.basename(p), open(p, "rb"))) for p in (second, first)]
    r = client.post("/uploads", files=payload)
    check("여러 파일 업로드 200", r.status_code == 200, r.text[:200])
    if r.status_code != 200:
        return

    d = r.json()
    check("파일별로 결과가 나뉨", len(d.get("uploads", [])) == 2,
          [u.get("filename") for u in d.get("uploads", [])])
    check("합계도 함께 옴", d["files"] == 2 and d["cards"] == 4,
          (d.get("files"), d.get("cards")))

    later = next(u for u in d["uploads"] if "2차수" in u["filename"])
    grew = 0
    for item in later["reports"]:
        if not item.get("report_id"):
            continue
        body = client.get(f"/reports/{item['report_id']}").json()["body"]
        if (body["sections"].get("growth") or {}).get("status") == "compared":
            grew += 1
    check("나중 차수를 먼저 올려도 성장 비교가 붙음", grew == 2, f"{grew}/2")


def case_http_roundtrip():
    """실제 업로드 경로(HTTP)로도 저장까지 통과하는지."""
    from fastapi.testclient import TestClient
    from main import app

    client = TestClient(app)
    path = os.path.join(TMP, "rich_meta.xlsx")
    with open(path, "rb") as fh:
        r = client.post("/uploads", files={"file": ("rich_meta.xlsx", fh)})
    check("업로드 HTTP 200", r.status_code == 200,
          r.text[:200] if r.status_code != 200 else "")
    if r.status_code == 200:
        d = r.json()
        made = [x for x in d.get("reports", []) if x.get("report_id")]
        check("리포트까지 생성", len(made) == 2, d.get("reports"))
        if made:
            h = client.get(made[0]["html"])
            check("리포트 HTML 응답", h.status_code == 200 and "<!DOCTYPE html>" in h.text)


# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    for fn in (case_rich_meta, case_no_scores, case_cover_sheet_first,
               case_merged_and_gaps, case_no_name_column,
               case_growth_mapping_drift, case_action_items_from_emphasis,
               case_gap_comment_rewrite, case_tone_normalisation,
               case_two_line_header_and_alias, case_english_comment_translation,
               case_emphasis_reinterpreted, case_blind_new_format,
               case_multi_file_upload, case_http_roundtrip):
        print(f"\n── {fn.__doc__.splitlines()[0]}")
        try:
            fn()
        except Exception as exc:                      # noqa: BLE001
            import traceback
            traceback.print_exc()
            _failures.append(f"{fn.__name__}: {type(exc).__name__}")

    print("\n" + ("전부 통과" if not _failures
                  else f"실패 {len(_failures)}건: {_failures}"))
    sys.exit(1 if _failures else 0)
